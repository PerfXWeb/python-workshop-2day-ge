
# This Python script helped me to create the fake dataset that we used for the excel project
import openpyxl
import os
import random
random.seed(1) # This sets the seed of the random module. Therefore, all random values will always be the same if we restart this program. Useful for development.

# Just a bunch of random names I got from a random name generator
# In total, this is 150 names
fake_names = [['Carl','Ince'], ['Madeleine','Young'], ['Wanda','Gibson'], ['Ryan','Poole'], ['Robert','Howard'], ['Stephen','Ellison'], ['Carol','Peters'], ['Joan','Ellison'], ['Dominic','Carr'], ['Ian','Martin'], ['Nicola','Parr'], ['Abigail','Langdon'], ['Oliver','Taylor'], ['Victoria','Roberts'], ['Jacob','Graham'], ['Brian','Cameron'], ['Jasmine','Ball'], ['Jonathan','Bower'], ['Evan','Gill'], ['Joan','Arnold'], ['Megan','Short'], ['Jane','Martin'], ['Andrea','Lee'], ['Peter','Duncan'], ['Hannah','Quinn'], ['Robert','Taylor'], ['Joe','Paterson'], ['Irene','Marshall'], ['Rebecca','Ogden'], ['Christopher','Randall'], ['Jan','Paige'], ['Ian','Gray'], ['Kimberly','Taylor'], ['Jonathan','Abraham'], ['Victor','Langdon'], ['Carl','Henderson'], ['Natalie','North'], ['Brian','Grant'], ['Molly','Walsh'], ['Anthony','Dowd'], ['Lucas','Quinn'], ['Lauren','White'], ['Dorothy','Murray'], ['Sally','Nash'], ['Ava','Jones'], ['Paul','Clark'], ['Jasmine','Nolan'], ['Dylan','Gill'], ['Ava','North'], ['Emma','Clark'], ['Zoe','Butler'], ['Liam','Roberts'], ['Connor','Johnston'], ['Justin','Alsop'], ['Carolyn','Powell'], ['James','Springer'], ['Samantha','MacLeod'], ['Victoria','Howard'], ['Gabrielle','Ince'], ['Julian','Welch'], ['Christopher','Hunter'], ['Nicola','Nash'], ['Ruth','Kerr'], ['Joan','Skinner'], ['Chloe','Henderson'], ['Steven','Forsyth'], ['Rose','Sanderson'], ['Piers','Newman'], ['Phil','Walsh'], ['Jack','Kerr'], ['Alison','Paige'], ['Christopher','Blake'], ['Anthony','Piper'], ['Joseph','Terry'], ['Diane','Cornish'], ['Deirdre','Abraham'], ['Megan','Alsop'], ['Jessica','Allan'], ['Alison','Mathis'], ['Vanessa','Allan'], ['Victor','Fraser'], ['Jake','Duncan'], ['Victoria','Grant'], ['Tim','Marshall'], ['Benjamin','Harris'], ['Penelope','McDonald'], ['Stephanie','Sutherland'], ['Sebastian','Anderson'], ['Oliver','Duncan'], ['Bella','Miller'], ['Megan','Hunter'], ['Nicholas','Lambert'], ['Gabrielle','Wilson'], ['Megan','Kerr'], ['Wendy','Paterson'], ['Pippa','Bell'], ['Tracey','Terry'], ['Isaac','Dyer'], ['Fiona','North'], ['Kimberly','Grant'], ['Diana','Sutherland'], ['Phil','Nash'], ['Megan','Peters'], ['Deirdre','Blake'], ['Gavin','Clark'], ['Sebastian','Rees'], ['Lily','Payne'], ['Blake','Gray'], ['Joanne','Ferguson'], ['Carol','Gill'], ['Max','Smith'], ['Isaac','Martin'], ['Molly','Abraham'], ['Joe','Piper'], ['Brandon','Campbell'], ['Evan','Anderson'], ['Jennifer','Churchill'], ['Alexander','Carr'], ['Leah','Rees'], ['Felicity','Underwood'], ['Robert','Wallace'], ['Alison','Wright'], ['Emma','Allan'], ['Rebecca','Rampling'], ['Jennifer','Abraham'], ['Brian','Chapman'], ['Isaac','Gibson'], ['Kimberly','Dickens'], ['Faith','Mitchell'], ['Lucas','Howard'], ['Max','Hamilton'], ['Donna','Baker'], ['Emily','Allan'], ['Eric','Slater'], ['Amelia','Turner'], ['Joan','Johnston'], ['Caroline','Cornish'], ['Charles','Bell'], ['Edward','Mackenzie'], ['Ella','Forsyth'], ['Stephen','Piper'], ['Victor','Vaughan'], ['Phil','James'], ['Austin','Bond'], ['Alexandra','Watson'], ['Alexandra','Cameron'], ['Julian','Wright'], ['Kevin','White'], ['Justin','Lyman'], ['Lillian','Pullman']]


if len(os.listdir('sheets')) < 10: # This if-sentence just checks for the number of files in the sheets directory so I don't create too many Excel sheets
    # opening the default Excel sheet
    wb = openpyxl.load_workbook(filename=os.path.join("sheets", "1 employee_template.xlsx"), data_only=True)
    ws = wb.active

    # now looping through all of those names (150 times)
    for x in fake_names:
        ws['B3'].value = x[0]+" "+x[1] # changing the name value in the sheet
        for i in range(6,58): # looping through all the cells for the individual weeks
            ws[f'B{i}'].value = random.randint(1,10) # adding some random value using random.randint(1,10) - basically any number from 1 to 10

        wb.save(filename = os.path.join("sheets", f"2 happiness_{x[0]}_{x[1]}.xlsx")) # lastly, saving the current file using a new name
else:
    print("Too many files in folder already")
