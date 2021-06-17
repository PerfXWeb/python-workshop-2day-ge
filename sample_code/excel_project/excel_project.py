

import os
import json
import glob
# need to pip-install these first
import requests # python3 -m pip install requests
import openpyxl # python3 -m pip install openpyxl


########## APPROACH ##########
#
# 1. Get the data out of every employee file
#   - To do so, we need to temporarily store the data in variables
#   - In the end, all we need is JUST the AVERAGE of the employees, not every single data point
#   - Technically, we could go through all 150 employee files 52 times and calculate the average of a single week every single time
#   - Instead, we will do the following (more efficient):
#       1. We will go though all employee files only ONCE and ADD UP the data points of each week
#       2. In the end, we have only 52 values, each one stores the SUM of all employees ratings of a single week
#       3. We can then divide this sum by the number of employee files to get the average
#
# 2. Calculate the average by using the SUM of the employee ratings divided by the number of employee files
# 3. Write the new averages into our main Excel sheet
# 4. Get COVID data from an API and insert into our Excel sheet
#   - We have to look at the API and check what data is actually being served
#   - The data from the API we are using provides us with TOTAL COVID cases of every single DAY
#   - Since we only need the data for every WEEK, we need to extract only every 7th data point
#   - We also need to make sure that the COVID data aligns with our week numbers.
#       - We are only provided with data starting at the end of February, so we need to adjust for that.


# Set up our list in which we will store all sums of our employee happiness for each week
sums = [0 for i in range(52)]

# Step 1: Get the data out of all employee sheets
# os.path.join - This helps us to correctly join folders and filenames in order for our code to work on all operating systems
# the asterisk "*" serves as a wildcard here, meaning that ANYTHING can be in between "2 happiness" and ".xlsx"
for filename in glob.glob(os.path.join('sheets', '2 happiness*.xlsx')):
    print(f"Opening: {filename}") # printing something always helps us to understand where our code is at the moment
    wb = openpyxl.load_workbook(filename=filename, data_only=True) # This opens our Excel file. "data_only=True" makes sure that we get the DATA from the sheet and not the formulas in a cell (if there are any)
    ws = wb.active # This opens the active (i.e. the first) worksheet in the Excel file
    for i in range(52): # The for loop runs 52 times to get 52 data points
        sums[i] = sums[i] + ws[f'B{i+6}'].value # We use ".value" to get the value of a specific cell. We also need to use {i+6} because we want to start getting the data from row number 6 (rather than row number 0 or 1, have a look at the Excel sheet)
        print(f"Week {i+1}: {ws[f'B{i+6}'].value}") # printing the current week (don't forget range() starts with ZERO) and the value of the cell


# ------------------------------------------------------------------------------------------------------------------------------


# Step 2: Calculate the averages
# Here we use "enumerate":
#   - enumerate() counts the number of times a loop has gone through. Since it also starts at 0, it is equal to the INDEX of the current value that we are looking at.
for e,i in enumerate(sums):
    sums[e] = i/len(glob.glob(os.path.join('sheets', '2 happiness*.xlsx'))) # We are changing every value in our list from being the SUM of all happiness values to being the AVERAGE of all happiness values

print("Step 2 done")


# ------------------------------------------------------------------------------------------------------------------------------


# Step 3: Write the new averages into that sheet
wb = openpyxl.load_workbook(filename=os.path.join("sheets", "0 correlation_sheet.xlsx"), data_only=True) # same as before
ws = wb.active # same as before
for e,i in enumerate(sums): # we need to use enumerate() again to know the current week we are looking at.
    print(f"Average for week {e+1}: {i}") # same as before, enumerate() starts at ZERO so we need to add 1 to also start with week 1
    ws[f'B{4+e}'].value = i # Once again, we need to use {4+e} because that is where our data should be inserted
wb.save(filename = os.path.join("sheets", "0 correlation_sheet.xlsx")) # Now we just need to save our new values


# ------------------------------------------------------------------------------------------------------------------------------


# Step 4: Get COVID data from an API and add it to the sheet
data = requests.get("https://api.covid19api.com/dayone/country/austria/status/confirmed") # This gets our raw JSON formatted API data
data = json.loads(data.content) # Since it is JSON data, we need to convert it to become a Python list that we can work with
wb = openpyxl.load_workbook(filename=os.path.join("sheets", "0 correlation_sheet.xlsx"), data_only=True) # same as before
ws = wb.active # same as before
for i in range(52-8): # Once again, we loop through every week. BUT we need to subtract 8 weeks since the COVID data is not provided for the first 8 weeks of 2020
    # Once again, we need to use {i+12} because that's where we first want to put our data in.
    # Since we only need the data for every 7 days, we are using [i*7] to multiply our current i value with 7. We also want to OFFSET our data gathering by 5 (to align with our week numbers)
    ws[f'C{i+12}'].value = data[i*7+5]['Cases'] # ['Cases'] holds the TOTAL number of cases so far
    ws[f'D{i+12}'].value = data[i*7+5]['Date'] # ['Date'] holds the date of that value. We use this to verify if we actually get the right data.

wb.save(filename = os.path.join("sheets", "0 correlation_sheet.xlsx")) # saving again, and done
